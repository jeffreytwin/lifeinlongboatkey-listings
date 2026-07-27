#!/usr/bin/env python3
"""Audit the Neighborhoods & Condos (NeighborhoodsCondos) Wix collection so every
'Nearby Neighborhood N - Description' equals the referenced neighborhood's
'Neighborhood Short Description' (field key: villageShortDescription).

Usage: python3 audit-nearby-descriptions.py <export.csv> <output.xlsx>

Input is the CSV exported from the Wix content manager. Output is an .xlsx with:
  - 'Paste into Wix'  : all rows in export order with corrected descriptions
                        (changed cells highlighted yellow)
  - 'Changed Cells'   : only the slots whose description changed, old vs new
  - 'Link Issues'     : nearby links whose slug doesn't match the referenced
                        neighborhood's current page slug
"""
import csv
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SLOTS = ['1', '2', '3', '4']


def norm_name(s):
    s = s.strip().lower()
    s = re.sub(r'^the\s+', '', s)
    s = s.replace('&', 'and')
    return re.sub(r'[^a-z0-9]+', ' ', s).strip()


def slug_of(url):
    m = re.search(r'/neighborhood(?:-homes-for-sale)?/([^/?#]+)', (url or '').strip())
    return m.group(1).lower() if m else ''


def build_resolver(rows):
    by_slug = {slug_of(r['Neighborhood Pages Main']): r for r in rows}
    by_name = {norm_name(r['Neighborhood Name']): r for r in rows}

    def resolve(title, link):
        s = slug_of(link)
        if s in by_slug:
            return by_slug[s], 'exact link'
        if s:
            alt = s[4:] if s.startswith('the-') else 'the-' + s
            if alt in by_slug:
                return by_slug[alt], 'link differs by "the-" prefix'
            n = norm_name(s.replace('-', ' '))
            if n in by_name:
                return by_name[n], 'link slug matched as name'
        if title and norm_name(title) in by_name:
            return by_name[norm_name(title)], 'matched by title (link slug not found)'
        if s:
            cands = [k for k in by_slug if k.startswith(s) or s.startswith(k)]
            if len(cands) == 1:
                return by_slug[cands[0]], 'link slug is a partial match'
        return None, None

    return resolve


def main(csv_path, out_path):
    with open(csv_path, encoding='utf-8-sig', newline='') as f:
        rows = list(csv.DictReader(f))
    resolve = build_resolver(rows)

    wb = Workbook()
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F3864')
    body_font = Font(name='Arial', size=10)
    changed_fill = PatternFill('solid', fgColor='FFF2CC')
    issue_fill = PatternFill('solid', fgColor='FCE4EC')
    wrap = Alignment(wrap_text=True, vertical='top')

    def style_sheet(ws, widths):
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical='center')
        ws.freeze_panes = 'A2'

    # Sheet 1: full grid in export order, descriptions replaced
    ws = wb.active
    ws.title = 'Paste into Wix'
    head = ['Neighborhood Name', 'ID']
    for i in SLOTS:
        head += [f'Nearby Neighborhood {i} - Title', f'Nearby Neighborhood {i} - Description']
    ws.append(head)

    changes, issues = [], []
    for r in rows:
        out = [r['Neighborhood Name'], r['ID']]
        marks = []
        for i in SLOTS:
            title = r[f'Nearby Neighborhood {i} - Title'].strip()
            link = r[f'Nearby Neighborhood {i} Link'].strip()
            cur = r[f'Nearby Neighborhood {i} - Description'].strip()
            ref, how = resolve(title, link)
            if ref is None:
                out += [title, cur]
                issues.append([r['Neighborhood Name'], i, title, link,
                               'UNRESOLVED - no neighborhood found; description left as-is', ''])
                continue
            new = ref['Neighborhood Short Description'].strip()
            out += [title, new]
            if new != cur:
                marks.append(len(out))
                changes.append([r['Neighborhood Name'], i, title,
                                ref['Neighborhood Name'], cur, new])
            if how != 'exact link':
                issues.append([r['Neighborhood Name'], i, title, link, how,
                               'https://www.lifeinlongboatkey.com' + ref['Neighborhood Pages Main']])
        ws.append(out)
        for cell in ws[ws.max_row]:
            cell.font = body_font
            cell.alignment = wrap
        for col in marks:
            ws.cell(row=ws.max_row, column=col).fill = changed_fill
    style_sheet(ws, [24, 38] + [22, 60] * 4)

    ws2 = wb.create_sheet('Changed Cells')
    ws2.append(['Neighborhood (row)', 'Slot', 'Nearby Title', 'Description pulled from',
                'Current Description', 'New Description (villageShortDescription)'])
    for row in changes:
        ws2.append(row)
        for cell in ws2[ws2.max_row]:
            cell.font = body_font
            cell.alignment = wrap
    style_sheet(ws2, [24, 6, 22, 24, 60, 60])

    ws3 = wb.create_sheet('Link Issues')
    ws3.append(['Neighborhood (row)', 'Slot', 'Nearby Title', 'Current Link',
                'Problem', 'Suggested Link'])
    for row in issues:
        ws3.append(row)
        for cell in ws3[ws3.max_row]:
            cell.font = body_font
            cell.alignment = wrap
            cell.fill = issue_fill
    style_sheet(ws3, [24, 6, 24, 55, 40, 55])

    wb.save(out_path)
    print(f'rows={len(rows)} changed={len(changes)} link_issues={len(issues)} -> {out_path}')


if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2])
